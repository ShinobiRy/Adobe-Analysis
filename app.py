from flask import Flask, render_template, request, send_file, jsonify
import os
import tempfile
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Configure upload folder and output file
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
    
OUTPUT_FILE = os.path.join(UPLOAD_FOLDER, 'combined_output.csv')
COLLEGE_STATS_FILE = os.path.join(UPLOAD_FOLDER, 'adobe_college_stats.xlsx')
ALLOWED_EXTENSIONS = {'csv', 'xlsx'}

# Configure a larger upload size limit
app.config['MAX_CONTENT_LENGTH'] = 1000 * 1024 * 1024  # 1GB max upload size

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Adobe Analysis Functions
def get_all_adobe_apps():
    """Return a comprehensive list of all Adobe applications to ensure complete reporting."""
    return [
        # Core Creative Suite
        "Adobe Photoshop",
        "Adobe Illustrator",
        "Adobe Premiere Pro",
        "Adobe After Effects",
        "Adobe InDesign",
        "Adobe Lightroom",
        "Adobe Acrobat",
        
        # Additional Design Apps
        "Adobe XD",
        "Adobe Dimension",
        "Adobe Animate",
        "Adobe Substance 3D",
        "Adobe Fresco",
        "Adobe Character Animator",
        "Adobe Express",
        
        # Video/Audio Apps
        "Adobe Audition",
        "Adobe Media Encoder",
        "Adobe SpeedGrade",
        "Adobe Prelude",
        
        # Publishing/Web Apps
        "Adobe Dreamweaver",
        "Adobe InCopy",
        "Adobe Bridge",
        "Adobe RoboHelp",
        
        # Other Adobe Apps
        "Adobe Scan",
        "Adobe Cloud Storage",
        "PDF Document",
        "Other Adobe Files"
    ]

def extract_adobe_app(item_path, debug=False):
    """
    Extract Adobe application name from the item path.
    If no known pattern is found, returns "Other Adobe Files".
    
    Parameters:
      item_path (str): The file path to evaluate.
      debug (bool): If True, print unprocessed paths for debugging.
    """
    if pd.isna(item_path):
        return "Unknown"
    
    item_path_str = str(item_path).lower()
    
    # Mapping of Adobe package identifiers to app names
    adobe_packages = {
        "com.adobe.acrobat": "Adobe Acrobat",
        "com.adobe.photoshop": "Adobe Photoshop",
        "com.adobe.illustrator": "Adobe Illustrator",
        "com.adobe.premiere": "Adobe Premiere Pro",
        "com.adobe.aftereffects": "Adobe After Effects",
        "com.adobe.lightroom": "Adobe Lightroom",
        "com.adobe.xd": "Adobe XD",
        "com.adobe.indesign": "Adobe InDesign",
        "com.adobe.animate": "Adobe Animate",
        "com.adobe.audition": "Adobe Audition",
        "com.adobe.dreamweaver": "Adobe Dreamweaver",
        "com.adobe.express": "Adobe Express"
    }
    
    # Check for app identifiers in the path
    for package, app_name in adobe_packages.items():
        if package in item_path_str:
            return app_name
            
    # Check for app names in path
    app_names = {
        "photoshop": "Adobe Photoshop",
        "illustrator": "Adobe Illustrator",
        "premiere": "Adobe Premiere Pro",
        "after effects": "Adobe After Effects",
        "lightroom": "Adobe Lightroom",
        "acrobat": "Adobe Acrobat",
        "xd": "Adobe XD",
        "indesign": "Adobe InDesign",
        "animate": "Adobe Animate",
        "express": "Adobe Express",
    }
    
    for app_pattern, app_name in app_names.items():
        if app_pattern in item_path_str:
            return app_name
    
    # Check for Lightroom-specific patterns (high priority)
    if any(pattern in item_path_str for pattern in ['/lightroom/', 'lightroom classic', '/lrcat/']):
        return "Adobe Lightroom"
    
    # Check for file extensions
    file_ext = os.path.splitext(item_path_str)[1]
    extension_apps = {
        # Core Creative Suite
        ".psd": "Adobe Photoshop",
        ".psdc": "Adobe Photoshop",
        ".psb": "Adobe Photoshop",
        ".aic": "Adobe Illustrator",
        ".ai": "Adobe Illustrator",
        ".prproj": "Adobe Premiere Pro",
        ".aep": "Adobe After Effects",
        ".express": "Adobe Express",
        ".indd": "Adobe InDesign",
        ".idrc": "Adobe InDesign",
        ".utxt": "Adobe InDesign",
        ".idml": "Adobe InDesign",  
        # Adobe Acrobat & PDFs
        ".acrobat": "Adobe Acrobat",  
        # Lightroom
        ".lrtemplate": "Adobe Lightroom",
        ".lrcat": "Adobe Lightroom",
        ".lrcat-wal": "Adobe Lightroom",
        ".lrcat-lock": "Adobe Lightroom",
        ".lrcat-shm": "Adobe Lightroom",
        ".lrprev": "Adobe Lightroom",
        # Additional Design Apps
        ".xd": "Adobe XD",
        ".xdc": "Adobe XD",
        ".dn": "Adobe Dimension",
        ".fla": "Adobe Animate",
        ".sbsar": "Adobe Substance 3D",
        ".fresco": "Adobe Fresco",
        ".chproj": "Adobe Character Animator",
        # Video/Audio Apps
        ".sesx": "Adobe Audition",
        ".prpreset": "Adobe Media Encoder",
        ".ircp": "Adobe SpeedGrade",
        ".plproj": "Adobe Prelude",
        # Publishing/Web Apps
        ".dw": "Adobe Dreamweaver",
        ".icml": "Adobe InCopy",
        ".brd": "Adobe Bridge"
    }
    
    if file_ext in extension_apps:
        return extension_apps[file_ext]
    
    # Special handling for PDFs
    if file_ext == '.pdf':
        if '/cloud-content/adobe scan/' in item_path_str:
            return 'Adobe Scan'
        return 'PDF Document'
    
    # System paths that indicate cloud storage
    if '/adobe-libraries/' in item_path_str or '/assets/adobe-libraries/' in item_path_str or '/cloud-content' in item_path_str:
        return 'Adobe Cloud Storage'
    
    # Optionally, log unprocessed paths for debugging purposes
    if debug:
        logger.debug(f"Unprocessed item path: {item_path_str}")
    
    return "Other Adobe Files"

def format_excel_sheet(worksheet):
    """
    Basic formatting for Excel worksheet
    """
    # Auto-adjust column width
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width
    
    # Format headers
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

def get_valid_colleges():
    """Return list of valid college abbreviations."""
    return [
        'ab', 'acct', 'archi', 'cfad', 'cics', 'comm', 'crs', 'cthm',
        'eccle', 'educ', 'ehs', 'eng', 'gs', 'gslaw', 'ipea', 'jhs',
        'law', 'med', 'music', 'nur', 'pharma', 'sc', 'sci', 'shs', 'gensan'
    ]

def extract_college_unit(email):
    """Extract college/unit from email address."""
    if not isinstance(email, str) or '@ust.edu.ph' not in email:
        return 'Non-UST'
        
    try:
        # Get the part before @ust.edu.ph
        username = email.split('@')[0]
        # Get the last part after the last dot
        parts = username.split('.')
        if len(parts) < 3:  # If email format is invalid
            return 'Others'
            
        college = parts[-1].lower()  # Get the last part (college abbreviation)
        valid_colleges = get_valid_colleges()
        
        if college in valid_colleges:
            return college.upper()
        else:
            return 'Others'
            
    except Exception:
        return 'Others'

def is_ust_student(email):
    """
    Determine if an email belongs to a UST student based on format:
    firstname.lastname.academicUnit@ust.edu.ph
    
    Returns True for valid UST student emails, False otherwise.
    """
    if not isinstance(email, str) or '@ust.edu.ph' not in email:
        return False
        
    try:
        # Get the part before @ust.edu.ph
        username = email.split('@')[0]
        # Get the parts split by dot
        parts = username.split('.')
        if len(parts) < 3:  # Must have at least firstname.lastname.unit
            return False
            
        # Get the academic unit (last part before @)
        academic_unit = parts[-1].lower()  
        valid_units = get_valid_colleges()
        
        # Check if it's a valid academic unit
        return academic_unit in valid_units
            
    except Exception:
        return False

def process_user_app_data(ust_df):
    """
    Process the first app usage for each student.
    Extracts and returns user-college-app mapping.
    """
    # Make a copy to avoid modifying the original dataframe
    df_copy = ust_df.copy()
    
    # Sort by timestamp if available to get first genuine app usage
    if 'Timestamp' in df_copy.columns:
        df_copy = df_copy.sort_values('Timestamp')
    
    # Use groupby to efficiently get first app usage per user
    first_usage = df_copy.groupby('User Email').first().reset_index()
    
    # Create result dataframe
    result_df = first_usage[['User Email', 'Adobe App', 'College']]
    
    return result_df

def get_highest_app_users(college_df, apps):
    """Extract highest users per app information."""
    highest_users = []
    
    # Filter out the "TOTAL" row first
    filtered_df = college_df[college_df['College'] != 'TOTAL']
    
    for app in apps:
        if app in filtered_df.columns:
            app_users = filtered_df[['College', app]]
            if not app_users.empty:
                max_users = app_users[app].max()
                if max_users > 0:
                    # Find colleges with this max value (could be multiple)
                    max_colleges = app_users[app_users[app] == max_users]['College']
                    if not max_colleges.empty:
                        college = max_colleges.iloc[0]
                        highest_users.append([app, college, max_users])
    
    # Sort highest to lowest by user count
    highest_users = sorted(highest_users, key=lambda x: x[2], reverse=True)
    return highest_users

def create_college_app_matrix(first_app_usage, valid_colleges, apps):
    """Create the college by app matrix for reporting."""
    college_app_matrix = []
    
    for college in valid_colleges:
        college_users = first_app_usage[first_app_usage['College'] == college]
        college_unique_users = len(college_users)
        row = [college, college_unique_users]  # College name and total unique users
        
        # Count for each app
        for app in apps:
            app_users = len(college_users[college_users['Adobe App'] == app])
            row.append(app_users)
        
        college_app_matrix.append(row)
    
    return college_app_matrix

def generate_college_usage_stats(df, output_file):
    """Generate Excel file with college-wise Adobe app usage statistics."""
    logger.info("Generating college usage statistics...")
    
    try:
        # Get row count metrics from the dataframe attributes
        total_rows = getattr(df, 'total_rows', len(df))
        duplicate_rows = getattr(df, 'duplicate_rows', 0)
        
        # Apply the new student classification
        df['is_ust_student'] = df['User Email'].apply(is_ust_student)
        
        # Ensure we're counting unique users correctly
        total_users = df['User Email'].nunique()
        ust_student_users = df[df['is_ust_student']]['User Email'].nunique()
        other_users = total_users - ust_student_users
        
        logger.info(f"Processing statistics for {total_users} unique users")
        logger.info(f"UST Student Users: {ust_student_users}, Other Users: {other_users}")
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: Overall Statistics - add row count metrics
            stats_df = pd.DataFrame({
                'Metric': [
                    'Total Unique Users', 
                    'UST Student Users', 
                    'Other Users',
                    'Total Rows',
                    'Duplicate Rows'
                ],
                'Count': [
                    total_users, 
                    ust_student_users, 
                    other_users,
                    total_rows,
                    duplicate_rows
                ]
            })
            
            stats_df.to_excel(writer, sheet_name='Overall Statistics', index=False)
            format_excel_sheet(writer.sheets['Overall Statistics'])
            
            # Sheet 2: College Distribution
            # Filter UST student emails and extract college
            ust_df = df[df['is_ust_student']].copy()
            ust_df['College'] = ust_df['User Email'].apply(extract_college_unit)
            
            # Process user app data
            first_app_usage = process_user_app_data(ust_df)
            
            # Get unique colleges and apps
            valid_colleges = [college.upper() for college in get_valid_colleges()]
            apps = get_all_adobe_apps()
            
            # Create cross-tabulation of colleges and apps
            college_app_matrix = create_college_app_matrix(first_app_usage, valid_colleges, apps)
            
            # Create column names
            columns = ['College', 'Total Unique Users'] + apps
            college_df = pd.DataFrame(college_app_matrix, columns=columns)
            
            # Add total row
            total_row = ['TOTAL', first_app_usage['College'].isin(valid_colleges).sum()]
            for app in apps:
                total_app_users = len(first_app_usage[
                    (first_app_usage['Adobe App'] == app) & 
                    first_app_usage['College'].isin(valid_colleges)
                ])
                total_row.append(total_app_users)
            
            college_df.loc[len(college_df)] = total_row
            
            # Get highest users per app
            highest_users = get_highest_app_users(college_df, apps)
            
            # Add empty row
            college_df.loc[len(college_df)] = [''] * len(college_df.columns)
            
            # Add header
            college_df.loc[len(college_df)] = ['Highest Users per App'] + [''] * (len(college_df.columns) - 1)
            
            # Add highest users data
            for app, college, count in highest_users:
                college_df.loc[len(college_df)] = [f'{app}:', college, count] + [''] * (len(college_df.columns) - 3)
            
            # Write main college distribution to Excel
            college_df.to_excel(writer, sheet_name='College Distribution', index=False)
            format_excel_sheet(writer.sheets['College Distribution'])
            
            # Sheet 3: Others (non-student users)
            other_users_df = df[~df['is_ust_student']].copy()
            if 'Timestamp' in other_users_df.columns:
                other_users_df = other_users_df.sort_values('Timestamp')
            
            # More efficient way to get first app usage
            if not other_users_df.empty:
                other_first_usage = other_users_df.groupby('User Email').first().reset_index()
                other_users_summary = other_first_usage[['User Email', 'Adobe App']].rename(
                    columns={'Adobe App': 'First Adobe App Used'})
                other_users_summary = other_users_summary.sort_values('User Email')
                
                # Write others to separate sheet
                other_users_summary.to_excel(writer, sheet_name='Other Users', index=False)
                format_excel_sheet(writer.sheets['Other Users'])
            
        # Format data for template display
        # Restructure highest_users to match template expectations
        highest_users_per_app = [(app, college, count) for app, college, count in highest_users if count > 0]
        
        # Only include entries that have both app and college
        highest_users_per_college = [(app, college) for app, college, _ in highest_users if college]
        
        # Filter college_df for actual colleges
        actual_colleges = college_df[
            (college_df['College'] != 'TOTAL') & 
            (college_df['College'] != '') & 
            (~college_df['College'].str.contains(':', na=False))
        ]
        
        # Only include valid colleges in the result
        valid_college_codes = set(valid_colleges)
        actual_colleges = actual_colleges[
            actual_colleges['College'].isin(valid_college_codes)
        ]
        
        actual_colleges_dict = actual_colleges[['College', 'Total Unique Users']].sort_values(
            'Total Unique Users', ascending=False).to_dict('records')
            
        preview_data = {
            "total_users": total_users,
            "ust_student_users": ust_student_users,
            "other_users": other_users,
            "total_rows": total_rows,
            "duplicate_rows": duplicate_rows,
            "all_colleges": actual_colleges_dict,
            "highest_users_per_app": sorted(highest_users_per_app, key=lambda x: x[2], reverse=True),
            "highest_users_per_college": sorted(highest_users_per_college, key=lambda x: x[1], reverse=True)
        }
            
        return True, preview_data
        
    except Exception as e:
        logger.error(f"Error generating college usage statistics: {str(e)}", exc_info=True)
        return False, {}
        
def process_files(files):
    """
    Process uploaded files and return a combined DataFrame with Adobe app detection.
    Ensures proper concatenation of all records and tracks duplicate rows.
    Also identifies users that appear in multiple files.
    """
    # Initialize tracking variables
    all_dataframes = []
    total_rows_before = 0
    users_per_file = {}
    
    # Process each file
    for i, file in enumerate(files):
        filename = secure_filename(file.filename)
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        try:
            # Save file to disk
            file.save(file_path)
            
            # Read the file based on its extension
            df = None
            if filename.endswith('.csv'):
                # Try different encodings
                encodings = ['utf-8', 'latin1', 'cp1252']
                for encoding in encodings:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding, low_memory=False)
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        logger.warning(f"Error reading {filename} with {encoding} encoding: {str(e)}")
                
                # If all encodings failed, try with Python engine as last resort
                if df is None:
                    df = pd.read_csv(file_path, encoding='utf-8', on_bad_lines='skip', engine='python')
            else:  # .xlsx
                df = pd.read_excel(file_path)
            
            # Track row counts and users
            total_rows_before += len(df)
            
            if 'User Email' in df.columns:
                # Convert to string first to handle NaN values safely
                df['User Email'] = df['User Email'].astype(str).str.strip()
                file_users = set(df['User Email'].dropna().unique())
                users_per_file[f"File {i+1}: {filename}"] = file_users
                logger.info(f"Processed {filename}: {len(df)} rows, {len(file_users)} unique users")
            else:
                logger.warning(f"Processed {filename}: {len(df)} rows, column 'User Email' not found")
            
            all_dataframes.append(df)
        except Exception as e:
            logger.error(f"Error processing file {filename}: {str(e)}", exc_info=True)
            raise ValueError(f"Error processing file {filename}: {str(e)}")
        finally:
            # Clean up the temp file
            if os.path.exists(file_path):
                os.remove(file_path)
    
    # Check if we have any valid dataframes
    if not all_dataframes:
        raise ValueError("No valid data found in uploaded files.")
    
    # Identify users that appear in multiple files
    logger.info("Identifying users appearing in multiple files")
    all_users_with_files = {}
    for file_name, users in users_per_file.items():
        for user in users:
            if user not in all_users_with_files:
                all_users_with_files[user] = []
            all_users_with_files[user].append(file_name)
    
    # Log users found in multiple files
    duplicate_users_across_files = {user: files for user, files in all_users_with_files.items() 
                                   if len(files) > 1 and user != 'nan'}
    if duplicate_users_across_files:
        logger.info(f"Found {len(duplicate_users_across_files)} users appearing in multiple files")
        for user, files in duplicate_users_across_files.items():
            logger.debug(f"User {user} appears in: {', '.join(files)}")
    else:
        logger.info("No users found in multiple files.")
    
    # Combine all dataframes
    combined_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
    
    # Validation: Make sure we didn't lose any rows
    total_rows_after = len(combined_df)
    logger.info(f"Total rows before concatenation: {total_rows_before}")
    logger.info(f"Total rows after concatenation: {total_rows_after}")
    
    if total_rows_before != total_rows_after:
        logger.warning("WARNING: Row count mismatch after concatenation!")
    
    # Verify that the required columns exist
    required_columns = ['User Email', 'Item Path']
    missing_columns = [col for col in required_columns if col not in combined_df.columns]
    
    if missing_columns:
        raise ValueError(f"Required columns missing in input files: {', '.join(missing_columns)}. "
                       f"Analysis requires columns: {', '.join(required_columns)}")
    
    # Clean data and ensure types are correct
    combined_df['User Email'] = combined_df['User Email'].astype(str).str.strip()
    combined_df['Item Path'] = combined_df['Item Path'].astype(str)
    
    # Count duplicates - rows where the same user accessed the same item path
    total_rows = len(combined_df)
    duplicate_count = total_rows - len(combined_df.drop_duplicates(subset=['User Email', 'Item Path']))
    logger.info(f"Total rows: {total_rows}, Duplicate rows: {duplicate_count}")
    
    # Process for Adobe analysis
    combined_df['Adobe App'] = combined_df['Item Path'].apply(extract_adobe_app)
    
    # Add row metrics to the dataframe as attributes
    combined_df.total_rows = total_rows
    combined_df.duplicate_rows = duplicate_count
    
    # Save combined file for reference
    combined_csv = os.path.join(UPLOAD_FOLDER, 'combined_adobe_logs.csv')
    combined_df.to_csv(combined_csv, index=False)
    
    return combined_df

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Check if files are in request
        if 'files' not in request.files:
            return render_template('index.html', error='No files selected')
        
        files = request.files.getlist('files')
        
        # Check if files were selected
        if not files or files[0].filename == '':
            return render_template('index.html', error='No files selected')
        
        # Filter only allowed files
        valid_files = [f for f in files if f and allowed_file(f.filename)]
        
        if not valid_files:
            return render_template('index.html', error='No valid CSV or XLSX files selected')
        
        try:
            # Process files - directly go to college stats generation
            combined_df = process_files(valid_files)
            success, preview_data = generate_college_usage_stats(combined_df, COLLEGE_STATS_FILE)
            
            if success:
                return render_template('index.html', success=True, file_count=len(valid_files), 
                                      preview_data=preview_data)
            else:
                return render_template('index.html', error="Failed to generate college statistics.")
                
        except ValueError as e:
            # Handle specific validation errors with clear messages
            logger.error(f"ValueError: {str(e)}", exc_info=True)
            return render_template('index.html', error=str(e))
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}", exc_info=True)
            
            # Create a more user-friendly error message
            error_msg = str(e)
            if "User Email" in error_msg:
                error_msg = "Analysis requires files with 'User Email' and 'Item Path' columns. Your files don't match this format."
                
            return render_template('index.html', error=f"Error processing files: {error_msg}")
    
    return render_template('index.html')

@app.route('/download')
def download():
    # Check if file exists first
    if not os.path.exists(COLLEGE_STATS_FILE):
        return render_template('index.html', error="No statistics file available for download. Please process files first.")
        
    # Direct download for college statistics file
    download_name = 'adobe_college_stats.xlsx'
    return send_file(COLLEGE_STATS_FILE, as_attachment=True, download_name=download_name)

# Add a custom error handler for the 413 entity too large error
@app.errorhandler(413)
def request_entity_too_large(error):
    return render_template('index.html', 
                          error="The file(s) you tried to upload are too large. Please keep total size under 1GB."), 413

if __name__ == '__main__':
    # Disable the reloader to avoid compatibility issues
    app.run(debug=True, use_reloader=False)